"""Import podmiotów do wniosku OK-WUD z 5 formatów: XLS, XLSX, CSV, ODS, XML.

Wszystkie szablony są w `docs/source/zus-bartosz-materials/Szablony - import/`.
Pola: LP, SYGNATURA_SPRAWY, NAZWA_SKROCONA, NAZWA_PELNA, IMIE, NAZWISKO,
NIP, REGON, PESEL, TYP_DOKUMENTU, NUMER_DOKUMENTU, DATA_URODZENIA,
KOD_POCZTOWY, MIEJSCOWOSC, ULICA, NR_DOMU, NR_LOKALU, PANSTWO.
"""
from __future__ import annotations

import csv
from datetime import date, datetime
from pathlib import Path
from typing import Any

from mcp_zus.okwud.models import PodmiotUdostepnioneDane

# Pola normalizowane (case-insensitive lookup)
_FIELD_MAP = {
    "lp": "lp",
    "sygnatura_sprawy": "sygnatura_sprawy",
    "nazwa_skrocona": "nazwa_skrocona",
    "nazwa_pelna": "nazwa_pelna",
    "imie": "imie",
    "nazwisko": "nazwisko",
    "nip": "nip",
    "regon": "regon",
    "pesel": "pesel",
    "typ_dokumentu": "rodzaj_dokumentu_tozsamosci",
    "numer_dokumentu": "numer_dokumentu_tozsamosci",
    "data_urodzenia": "data_urodzenia",
    "kod_pocztowy": "kod_pocztowy",
    "miejscowosc": "miejscowosc",
    "ulica": "ulica",
    "nr_domu": "nr_domu",
    "nr_lokalu": "nr_lokalu",
    "panstwo": "panstwo",
}


def _normalize_key(key: str) -> str | None:
    return _FIELD_MAP.get(key.strip().lower().replace(" ", "_"))


def _row_to_podmiot(row: dict[str, Any]) -> PodmiotUdostepnioneDane:
    """Convert one row (dict of source-key → value) to PodmiotUdostepnioneDane."""
    payload: dict[str, Any] = {}
    for raw_key, raw_val in row.items():
        if raw_val is None or raw_val == "":
            continue
        nk = _normalize_key(raw_key)
        if nk is None:
            continue
        if nk == "rodzaj_dokumentu_tozsamosci":
            try:
                payload[nk] = int(raw_val)
            except (TypeError, ValueError):
                continue
        elif nk == "data_urodzenia":
            payload[nk] = _parse_date(raw_val)
        elif nk == "lp":
            continue  # numer porządkowy — pomijamy
        else:
            payload[nk] = str(raw_val).strip()
    return PodmiotUdostepnioneDane(**payload)


def _parse_date(value: Any) -> date | None:
    if isinstance(value, date) and not isinstance(value, datetime):
        return value
    if isinstance(value, datetime):
        return value.date()
    if not value:
        return None
    s = str(value).strip()
    for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%d/%m/%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def import_from_csv(path: str | Path) -> list[PodmiotUdostepnioneDane]:
    path = Path(path)
    out: list[PodmiotUdostepnioneDane] = []
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        # Wykryj separator: ZUS używa ; ale ludzie wrzucają też ,
        sample = f.read(2048)
        f.seek(0)
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=";,\t")
        except csv.Error:
            dialect = csv.excel
            dialect.delimiter = ";"
        reader = csv.DictReader(f, dialect=dialect)
        for row in reader:
            out.append(_row_to_podmiot(row))
    return out


def import_from_xlsx(path: str | Path) -> list[PodmiotUdostepnioneDane]:
    from openpyxl import load_workbook

    wb = load_workbook(filename=str(path), read_only=True, data_only=True)
    ws = wb.active
    if ws is None:
        return []
    rows_iter = ws.iter_rows(values_only=True)
    headers = [str(c).strip() if c is not None else "" for c in next(rows_iter)]
    out: list[PodmiotUdostepnioneDane] = []
    for row in rows_iter:
        row_dict = dict(zip(headers, row, strict=False))
        if all(v in (None, "") for v in row_dict.values()):
            continue
        out.append(_row_to_podmiot(row_dict))
    return out


def import_from_xls(path: str | Path) -> list[PodmiotUdostepnioneDane]:
    import xlrd

    book = xlrd.open_workbook(str(path))
    sheet = book.sheet_by_index(0)
    if sheet.nrows == 0:
        return []
    headers = [str(sheet.cell_value(0, c)).strip() for c in range(sheet.ncols)]
    out: list[PodmiotUdostepnioneDane] = []
    for r in range(1, sheet.nrows):
        row_dict = {headers[c]: sheet.cell_value(r, c) for c in range(sheet.ncols)}
        if all(v in (None, "") for v in row_dict.values()):
            continue
        out.append(_row_to_podmiot(row_dict))
    return out


def import_from_ods(path: str | Path) -> list[PodmiotUdostepnioneDane]:
    from odf.opendocument import load
    from odf.table import Table, TableCell, TableRow
    from odf.text import P

    doc = load(str(path))
    tables = doc.spreadsheet.getElementsByType(Table)
    if not tables:
        return []
    rows = tables[0].getElementsByType(TableRow)
    if not rows:
        return []

    def _cell_text(cell: Any) -> str:
        return "".join(str(p) for p in cell.getElementsByType(P))

    headers = [_cell_text(c).strip() for c in rows[0].getElementsByType(TableCell)]
    out: list[PodmiotUdostepnioneDane] = []
    for row in rows[1:]:
        cells = row.getElementsByType(TableCell)
        values = [_cell_text(c) for c in cells]
        row_dict = dict(zip(headers, values, strict=False))
        if all(v == "" for v in row_dict.values()):
            continue
        out.append(_row_to_podmiot(row_dict))
    return out


def import_from_xml(path: str | Path) -> list[PodmiotUdostepnioneDane]:
    from lxml import etree

    tree = etree.parse(str(path))
    out: list[PodmiotUdostepnioneDane] = []
    for podm in tree.iter():
        tag = etree.QName(podm).localname.lower()
        if tag in {"podmiot", "subject", "wiersz", "row"}:
            row_dict: dict[str, Any] = {}
            for child in podm:
                ctag = etree.QName(child).localname.lower()
                row_dict[ctag] = child.text
            if any(row_dict.values()):
                out.append(_row_to_podmiot(row_dict))
    return out


def import_subjects(path: str | Path) -> list[PodmiotUdostepnioneDane]:
    """Auto-detect format by extension and import."""
    path = Path(path)
    suffix = path.suffix.lower()
    match suffix:
        case ".csv":
            return import_from_csv(path)
        case ".xlsx":
            return import_from_xlsx(path)
        case ".xls":
            return import_from_xls(path)
        case ".ods":
            return import_from_ods(path)
        case ".xml":
            return import_from_xml(path)
        case _:
            raise ValueError(f"Unsupported format: {suffix}. Use CSV/XLSX/XLS/ODS/XML.")


__all__ = [
    "import_from_csv",
    "import_from_ods",
    "import_from_xls",
    "import_from_xlsx",
    "import_from_xml",
    "import_subjects",
]
